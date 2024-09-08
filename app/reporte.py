import openpyxl

class GeneradorReportes:
    def __init__(self, citas):
        self.citas = citas  
        self.formatos = []

    def agregar_formato(self, formato):
        self.formatos.append(formato)

    def generarReporte(self):
        for formato in self.formatos:
            formato.generar(self.citas) 

class FormatoExcel:
    def generar(self, citas):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 'Reporte de Citas'
        ws['A2'] = 'ID Cita'
        ws['B2'] = 'Paciente'
        ws['C2'] = 'Médico'
        ws['D2'] = 'Fecha'
        ws['E2'] = 'Hora'

       
        print(f"{'ID Cita':<10}{'Paciente':<25}{'Médico':<25}{'Fecha':<15}{'Hora':<10}")
        print("="*85)

        
        for cita in citas:
            row = [cita.idCita, cita.paciente.nombre_completo, cita.medico.nombre_completo, cita.fecha, cita.hora]
            ws.append(row)

           
            print(f"{row[0]:<10}{row[1]:<25}{row[2]:<25}{row[3]:<15}{row[4]:<10}")

        
        archivo = "ReporteCitas.xlsx"
        try:
            wb.save(archivo)
            print(f"Reporte Excel generado con éxito en {archivo}.")
        except Exception as e:
            print(f"Error al generar el reporte: {e}")


if __name__ == "__main__":
   
    citas_prueba = []
    
    generador = GeneradorReportes(citas_prueba)
    formato_excel = FormatoExcel()
    generador.agregar_formato(formato_excel)
    generador.generarReporte()
